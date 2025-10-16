import csv, io
from typing import Any, Dict, List, Mapping, Optional, Sequence, cast
from fastapi import Response

def _normalize_rows(rows: Optional[Sequence[Mapping[str, Any]]]) -> List[Dict[str, Any]]:
    if not rows:
        return [dict()]
    return [dict(r) for r in rows]

def write_csv(rows: Optional[Sequence[Mapping[str, Any]]], filename: str) -> Response:
    data = _normalize_rows(rows)
    headers = sorted({k for r in data for k in r.keys()})
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=headers)
    w.writeheader()
    for r in data:
        w.writerow({k: r.get(k, "") for k in headers})
    return Response(
        buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def write_xlsx(rows: Optional[Sequence[Mapping[str, Any]]], filename: str) -> Response:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    data = _normalize_rows(rows)
    headers = sorted({k for r in data for k in r.keys()})

    wb = openpyxl.Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "export"

    if headers:
        ws.append(headers)
    for r in data:
        ws.append([r.get(h, "") for h in headers])

    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(40, len(str(h)) + 2))

    bio = io.BytesIO()
    wb.save(bio)
    return Response(
        bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
